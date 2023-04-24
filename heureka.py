#%%
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows


# read in the Excel file and create a DataFrame
heureka_stand_source = pd.read_csv('inputs/sweden/heureka/stand_from_heureka/stand_ulf.csv', sep=';')
output = 'inputs/sweden/heureka/'
species_codes ={'Pine':1, 'Spruce':2, 'Birch':3}


# %%

for site in heureka_stand_source.Description.unique():
    
# if True == True:
#     sites = heureka_stand_source.Description.unique()
#     site = sites[0]
#     print(site)
    
    heureka_stand = None
    #heureka_stand = heureka_stand_source[heureka_stand_source['Description'] == site]
    heureka_stand = heureka_stand_source.loc[heureka_stand_source['Description'] == site]
    heureka_stand = heureka_stand.reset_index(drop=True)

    # create StandData DataFrame
    StandData = pd.DataFrame()
    StandData['Simulation'] = [1] * len(heureka_stand)
    StandData['Year'] = heureka_stand['PERIOD']
    StandData['Age'] = heureka_stand['ForestData Mean Age (excl overstorey)'].astype(int)
    StandData['N'] = heureka_stand['ForestData Stems']
    StandData['BA'] = heureka_stand['ForestData Basal area (excl overstorey)']
    StandData['Hg'] = heureka_stand['ForestData Hgv']
    StandData['Dg'] = heureka_stand['ForestData Dgv']
    StandData['hdom'] = heureka_stand['ForestData DominantHeight']
    StandData['Total_volume'] = heureka_stand['ForestData Volume (excl overstorey)']
    StandData['Log'] = heureka_stand['BiomassData Biomass Stems All Species']
    StandData['Pulpwood'] = 0
    StandData['Waste_wood'] = 0
    StandData['Yield'] = 0
    StandData['Mortality'] = 0
    StandData['stem(commercial wood)'] = 0
    StandData['stem(waste)'] = 0
    StandData['living branches'] = heureka_stand['BiomassData Biomass Branches All Species']
    StandData['dead branches'] = heureka_stand['BiomassData Biomass Dead Branches All Species']
    StandData['foliage'] = heureka_stand['BiomassData Biomass Foliage All Species']
    StandData['Stumps'] = 0
    StandData['Roots >2mm'] = heureka_stand['BiomassData Biomass Stump and Roots ≥ 5 mm of All Species']
    StandData['Fine roots'] = heureka_stand['BiomassData Biomass Roots 2-5 mm of All Species']

    print(StandData.head(5))

    # create Removal DataFrame
    Removal = pd.DataFrame()
    Removal['Simulation'] = [1]
    Removal['Year'] = [heureka_stand['PERIOD'].max()]
    Removal['id Thinning'] = 0
    Removal['Thinning'] = 'Final cut'
    Removal['id Species'] = species_codes.get(heureka_stand['ForestData Dominant Species'].iloc[0], 0)
    Removal['Species'] = heureka_stand['ForestData Dominant Species'].iloc[0]
    Removal['Log[m³/ha]'] = 0
    Removal['Small log[m³/ha]'] = 0
    Removal['Pulpwood[m³/ha]'] = 0
    Removal['Energy wood, stem(commercial wood)[m³/ha]'] = 0
    Removal['Energy wood, stem(waste)[m³/ha]'] = 0
    Removal['Energy wood, branches(*->e)[m³/ha]'] = 0
    Removal['Energy wood, branches(*->k)[m³/ha]'] = 0
    Removal['Energy wood, *->kannot ja juuret[m³/ha]'] = 0

    # write file  
    workbook = openpyxl.Workbook()
    stand_sheet = workbook.create_sheet('StandData',0)
    for row in dataframe_to_rows(StandData, index=False, header=True):
        stand_sheet.append(row)

    # Add Removal sheet
    removal_sheet = workbook.create_sheet('Removal',1)
    for row in dataframe_to_rows(Removal, index=False, header=True):
        removal_sheet.append(row)

    # Save workbook
    file = output + site + '_heureka_input_lyr_0.xlsx'
    workbook.save(file)


    print(file)

# %%
print(workingFolder)

# read weather input
forc=read_FMI_weather(0, start_date, end_date, sourcefile=wfile)

# %%
